##region Setttings
sInputFolderPath = "../res/Input"
bPause = True
bSkipScrapping = False
sOutputLogName = "__Log.txt"
##endregion
##region Imports
import os
import FormatRosterData as FRD
import openpyxl
import TM_CommonPy as TM
import traceback
from FormatRosterData._Logger import FRDLog
from FormatRosterData._Logger import vFormatter #awkward
from FormatRosterData._Reporter import Reporter
import logging
##endregion

def AddHandlerToFRDLog_OutputLog():
    vFileHandler = logging.FileHandler(sOutputLogName)
    vFileHandler.setFormatter(vFormatter)
    vFileHandler.setLevel(1)
    FRDLog.addHandler(vFileHandler)

def Main():
    with TM.WorkspaceContext("Output",bCDInto=True,bPreDelete=True):
        AddHandlerToFRDLog_OutputLog()
        vReporter = Reporter()
        #-
        cWorkbooksToReformat = [] #List<Tuple<vOldWorkbook,sFileName>>
        c2dNameToUrl = {} #Dict<eSportCatagory:cNameToURL>
        #---Collect and dump c2dNameToUrl
        FRDLog.info("---Getting NameToURL lists---")
        #-Scrape known NameToURL websites
        for eSportCatagory in FRD.SportCatagory:
            c2dNameToUrl[eSportCatagory] = FRD.ScrapeNameToURLDict(eSportCatagory)
        #-in Input folder, parse any NameToURL txt files
        for sFileName in os.listdir(sInputFolderPath):
            sFileNameExtension = sFileName.split(".")[-1]
            if sFileNameExtension == "txt" and "nametourl" in sFileName.lower():
                #-Determine eSportCatagory
                eSportCatagory = FRD.SportCatagory.GetFromString(sFileName)
                if eSportCatagory is None:
                    FRDLog.warning("NameToURL`Could not determine eSportCatagory from:"+sFileName)
                    vReporter.cNameToURLErrorMsgs.append(sFileName)
                    continue
                #-Parse
                with open(os.path.join(sInputFolderPath,sFileName),'r') as vTextFile:
                    for sLine in vTextFile.readlines():
                        sLine = TM.RemoveWhitespace(sLine)
                        if not sLine:
                            continue
                        cKeyValuePair = sLine.split(":",1)
                        if len(cKeyValuePair) != 2:
                            FRDLog.warning("NameToURL`Incorrectly formatted line:"+sLine)
                            vReporter.cNameToURLErrorMsgs.append(sFileName+")Incorrectly formatted line:"+sLine.replace("\n",""))
                            continue
                        if cKeyValuePair[0] in c2dNameToUrl[eSportCatagory]:
                            sMsg = "NameToURL`Name already exists in c2dNameToUrl["+SportCatagory(eSportCatagory).name+"]:"+sLine
                            FRDLog.warning(sMsg)
                            vReporter.cNameToURLErrorMsgs.append(sMsg)
                            continue
                        #-
                        FRDLog.debug("NameToURL`Adding key:"+cKeyValuePair[0]+" value:"+cKeyValuePair[1])
                        c2dNameToUrl[eSportCatagory][cKeyValuePair[0]] = cKeyValuePair[1]
        #-Write a cNameToURL.txt for each catagory
        for eSportCatagory in FRD.SportCatagory:
            sOutputNameToURLFileName = '__cNameToURL_'+FRD.SportCatagory(eSportCatagory).name+'.txt'
            FRD.WriteDictToTxtFile(c2dNameToUrl[eSportCatagory],sOutputNameToURLFileName)
            FRDLog.info(sOutputNameToURLFileName+" complete")
        #---Get OldWorkbooks
        FRDLog.info("---Collecting unformatted sheets---")
        for sFileName in os.listdir(sInputFolderPath):
            sFileNameExtension = sFileName.split(".")[-1]
            if (not sFileNameExtension in ["xlsx","txt"]) or "~$" in sFileName or "template" in sFileName.lower():
                FRDLog.info("IGNORED:"+sFileName)
                continue
            elif sFileNameExtension == "txt" and "URL" in sFileName and "men" in sFileName.lower():
                pass
            elif sFileNameExtension == "xlsx":
                vOldWorkbook = openpyxl.load_workbook(os.path.join(sInputFolderPath,sFileName))
                cWorkbooksToReformat.append((vOldWorkbook,sFileName))
            elif sFileNameExtension == "txt" and not bSkipScrapping:
                #-Determine eSportCatagory
                eSportCatagory = FRD.SportCatagory.GetFromString(sFileName)
                if eSportCatagory is None:
                    FRDLog.warning("GetOldWorkbooks`Could not determine eSportCatagory from:"+sFileName)
                    vReporter.cGetOldWorkbookMsgs.append(sFileName)
                    continue
                #-Try to match names in text file to c2dNameToUrl[eSportCatagory].
                with open(os.path.join(sInputFolderPath,sFileName),'r') as vTextFile:
                    for sLine in vTextFile.readlines():
                        if not sLine:
                            continue
                        sLine = sLine.rstrip('\n').lower()
                        for vKey in c2dNameToUrl[eSportCatagory].keys():
                            if sLine in vKey.lower():
                                FRDLog.info(sFileName+" -  MATCHED:"+sLine+"("+vKey+")")
                                if "women" in sFileName.lower():
                                    sScrapedFileName = "Scraped_Women_"+FRD.GetTitle(c2dNameToUrl[eSportCatagory][vKey])
                                else:
                                    sScrapedFileName = "Scraped_Men_"+FRD.GetTitle(c2dNameToUrl[eSportCatagory][vKey])
                                vOldWorkbook = FRD.GetWorkbook(c2dNameToUrl[eSportCatagory][vKey])
                                cWorkbooksToReformat.append((vOldWorkbook,sScrapedFileName))
                                break
                        else:
                            vReporter.cUnmatchedFileNames.append(sLine)
                            FRDLog.info(sFileName+" - Could not match:"+sLine)
            else:
                vReporter.cUnmatchedFileNames.append(sFileName)
                FRDLog.warning("Could not get workbook from file:"+sFileName)
                continue
        #---Create NewWorkbooks
        FRDLog.info("---Creating formatted sheets---")
        for vOldWorkbook, sFileName in cWorkbooksToReformat:
            FRDLog.warning.ClearCallCount()
            FRDLog.info("OldFileName:"+sFileName)
            #---Edit
            vOldSheet = vOldWorkbook.active
            vNewWorkbook = openpyxl.Workbook()
            vNewSheet = vNewWorkbook.active
            FRD.FormatName(vOldSheet,vNewSheet)
            FRD.FormatHometown(vOldSheet,vNewSheet)
            FRD.FormatHeight(vOldSheet,vNewSheet)
            if not "women" in sFileName.lower():
                FRD.FormatWeight(vOldSheet,vNewSheet)
            FRD.FormatSchoolyear(vOldSheet,vNewSheet)
            FRD.FormatPosition(vOldSheet,vNewSheet)
            FRD.AppendOldSheet(vOldSheet,vNewSheet)
            #---Save
            if FRDLog.warning.iCallCount:
                sFileName = "_ERRORS_"+sFileName.split(".")[0]+"_Reformatted.xlsx"
                vReporter.cFormattingErrorFileNames.append(sFileName)
            else:
                sFileName = sFileName.split(".")[0]+"_Reformatted.xlsx"
                vReporter.cSuccessfulFormatFileNames.append(sFileName)
            FRDLog.info("New_FileName:"+sFileName)
            vNewWorkbook.save(sFileName)
        #---Report
        vReporter.WriteReport()
        #-Remove OutputLog handler from FRDLog
        FRDLog.handlers = [h for h in FRDLog.handlers if hasattr(h,"baseFilename") and sOutputLogName != os.path.basename(h.baseFilename)]


try:
    Main()
except PermissionError:
    FRDLog.error("\n\tI'd recommend to just try again.\n\tOtherwise, close all extra programs and then retry.",extra={'sOverrideLevelName': "PERMISSION_ERROR"})
    TM.DisplayDone()
except Exception as e:
    TM.DisplayException(e)
else:
    if bPause:
        TM.DisplayDone()
