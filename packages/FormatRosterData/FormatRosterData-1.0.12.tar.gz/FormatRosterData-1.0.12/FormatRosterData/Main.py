##region Setttings
sInputFolderPath = "../res/Input"
bPause = True
bSkipScrapping = False
##endregion
##region Imports
import os
import FormatRosterData as FRD
import openpyxl
import TM_CommonPy as TM
import traceback
from FormatRosterData._Logger import FRDLog
from FormatRosterData._Logger import vFormatter #awkward
import logging
##endregion

def Main():
    with TM.WorkspaceContext("Output",bCDInto=True,bPreDelete=True):
        #-Add Report_Full.txt handler to FRDLog
        vFileHandler = logging.FileHandler("__Report_Full.txt")
        vFileHandler.setFormatter(vFormatter)
        vFileHandler.setLevel(1)
        FRDLog.addHandler(vFileHandler)
        #-
        cUnmatchedFileNames = []
        cFormattingErrorFileNames = []
        cSuccessfulFormatFileNames = []
        cWorkbooksToReformat = [] #Expects value to be tuple(vOldWorkbook,sFileName)
        #---Output cNameToURL_Men.txt, cNameToURL_Women.txt
        FRDLog.info("---Getting NameToURL lists---")
        cNameToURL_Men = FRD.GetDict_NameToURL_Men()
        cNameToURL_Women = FRD.GetDict_NameToURL_Women()
        for sFileName in os.listdir(sInputFolderPath):
            sFileNameExtension = sFileName.split(".")[-1]
            if sFileNameExtension == "txt" and "URL" in sFileName and "men" in sFileName.lower():
                with open(os.path.join(sInputFolderPath,sFileName),'r') as vTextFile:
                    for sLine in vTextFile.readlines():
                        cStrings = TM.RemoveWhitespace(sLine).split(":",1)
                        if len(cStrings) != 2:
                            FRDLog.warning("NameToURL`Incorrectly formatted line:"+sLine)
                            continue
                        #-
                        if "women" in sFileName.lower():
                            if cStrings[0] in cNameToURL_Women:
                                FRDLog.warning("NameToURL`Name already exists in cNameToURL_Women:"+sLine)
                            else:
                                FRDLog.debug("NameToURL`Women. key:"+cStrings[0]+" value:"+cStrings[1])
                                cNameToURL_Women[cStrings[0]] = cStrings[1]
                        else:
                            if cStrings[0] in cNameToURL_Men:
                                FRDLog.warning("NameToURL`Name already exists in cNameToURL_Men:"+sLine)
                            else:
                                FRDLog.debug("NameToURL`Men. key:"+cStrings[0]+" value:"+cStrings[1])
                                cNameToURL_Men[cStrings[0]] = cStrings[1]
        FRD.WriteDictToTxtFile(cNameToURL_Women,'__cNameToURL_Women.txt')
        FRDLog.info("__cNameToURL_Women.txt complete")
        FRD.WriteDictToTxtFile(cNameToURL_Men,'__cNameToURL_Men.txt')
        FRDLog.info("__cNameToURL_Men.txt complete")
        #---Get OldWorkbooks
        FRDLog.info("---Collecting unformatted sheets---")
        for sFileName in os.listdir(sInputFolderPath):
            sFileNameExtension = sFileName.split(".")[-1]
            if (not sFileNameExtension in ["xlsx","txt"]) or "~$" in sFileName or "template" in sFileName.lower():
                FRDLog.info("IGNORED:"+sFileName)
                continue
            elif sFileNameExtension == "txt" and "URL" in sFileName and "men" in sFileName.lower():
                pass
            elif sFileNameExtension == "xlsx":
                vOldWorkbook = openpyxl.load_workbook(os.path.join(sInputFolderPath,sFileName))
                cWorkbooksToReformat.append((vOldWorkbook,sFileName))
            elif sFileNameExtension == "txt" and not bSkipScrapping:
                #-Determine which cNameToURL to use
                if "women" in sFileName.lower():
                    cNameToURL = cNameToURL_Women
                else:
                    cNameToURL = cNameToURL_Men
                #-Try to match names in text file to cNameToURL.
                with open(os.path.join(sInputFolderPath,sFileName),'r') as vTextFile:
                    for sLine in vTextFile.readlines():
                        if not sLine:
                            continue
                        sLine = sLine.rstrip('\n').lower()
                        for vKey in cNameToURL.keys():
                            if sLine in vKey.lower():
                                FRDLog.info(sFileName+" -  MATCHED:"+sLine+"("+vKey+")")
                                if "women" in sFileName.lower():
                                    sScrapedFileName = "Scraped_Women_"+FRD.GetTitle(cNameToURL[vKey])
                                else:
                                    sScrapedFileName = "Scraped_Men_"+FRD.GetTitle(cNameToURL[vKey])
                                vOldWorkbook = FRD.GetWorkbook(cNameToURL[vKey])
                                cWorkbooksToReformat.append((vOldWorkbook,sScrapedFileName))
                                break
                        else:
                            cUnmatchedFileNames.append(sLine)
                            FRDLog.info(sFileName+" - Could not match:"+sLine)
            else:
                cUnmatchedFileNames.append(sFileName)
                FRDLog.warning("Could not get workbook from file:"+sFileName)
                continue
        #---Create NewWorkbooks
        FRDLog.info("---Creating formatted sheets---")
        for vOldWorkbook, sFileName in cWorkbooksToReformat:
            FRDLog.warning.ClearCallCount()
            FRDLog.info("OldFileName:"+sFileName)
            #---Edit
            vOldSheet = vOldWorkbook.active
            vNewWorkbook = openpyxl.Workbook()
            vNewSheet = vNewWorkbook.active
            FRD.FormatName(vOldSheet,vNewSheet)
            FRD.FormatHometown(vOldSheet,vNewSheet)
            FRD.FormatHeight(vOldSheet,vNewSheet)
            if not "women" in sFileName.lower():
                FRD.FormatWeight(vOldSheet,vNewSheet)
            FRD.FormatSchoolyear(vOldSheet,vNewSheet)
            FRD.FormatPosition(vOldSheet,vNewSheet)
            FRD.AppendOldSheet(vOldSheet,vNewSheet)
            #---Save
            if FRDLog.warning.iCallCount:
                sFileName = "_ERRORS_"+sFileName.split(".")[0]+"_Reformatted.xlsx"
                cFormattingErrorFileNames.append(sFileName)
            else:
                sFileName = sFileName.split(".")[0]+"_Reformatted.xlsx"
                cSuccessfulFormatFileNames.append(sFileName)
            FRDLog.info("New_FileName:"+sFileName)
            vNewWorkbook.save(sFileName)
        #---Report
        FRDLog.info("---Report---")
        #-
        if cSuccessfulFormatFileNames:
            FRDLog.info("\t"+str(len(cSuccessfulFormatFileNames)) + " SUCCESSFULLY FORMATTED FILES")
            for sFileName in cSuccessfulFormatFileNames:
                FRDLog.info(sFileName)
        else:
            FRDLog.info("There were no successfully formatted files.")
        #-Add Report.txt handler to FRDLog
        vFileHandler = logging.FileHandler("__Report_Unmatched,ErrorFiles.txt")
        vFileHandler.setFormatter(logging.Formatter('%(message)s'))
        vFileHandler.setLevel(1)
        FRDLog.addHandler(vFileHandler)
        #-
        if cFormattingErrorFileNames:
            FRDLog.info("\t"+str(len(cFormattingErrorFileNames)) + " ERROR FILES")
            for sFileName in cFormattingErrorFileNames:
                FRDLog.info(sFileName)
        else:
            FRDLog.info("There were no errors.")
        if cUnmatchedFileNames:
            FRDLog.info("\t"+str(len(cUnmatchedFileNames)) + " UNMATCHED")
            for sFileName in cUnmatchedFileNames:
                FRDLog.info(sFileName)
        else:
            FRDLog.info("There were no unmatched files.")
        #-Remove 'Report' handlers from FRDLog
        FRDLog.handlers = [h for h in FRDLog.handlers if hasattr(h,"baseFilename") and not "report" in os.path.basename(h.baseFilename).lower()]


try:
    Main()
except PermissionError:
    FRDLog.error("\n\tI'd recommend to just try again.\n\tOtherwise, close all extra programs and then retry.",extra={'sOverrideLevelName': "PERMISSION_ERROR"})
    TM.DisplayDone()
except Exception as e:
    TM.DisplayException(e)
else:
    if bPause:
        TM.DisplayDone()
