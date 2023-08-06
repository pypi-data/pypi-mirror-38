#    This file is part of zBrac.
#
#    zBrac is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    zBrac is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with zBrac.  If not, see <https://www.gnu.org/licenses/>.

import sys
import os
import re
import xlsxwriter
import openpyxl
import sys
import os

def load_treatment_file(filepath, encoding = "UTF-8"):
    #with open(filepath, encoding="ISO-8859-1") as file: # German
    try:
        with open(filepath, encoding=encoding) as file: # Turkish
            filedata = file.read()
        file.close()
    except:
        print('! Warning: Non readable charachters exist for encoding ' + encoding + '. \n Ignoring non convertible chars.\n Please check if you have selected right encoding')
        with open(filepath, encoding=encoding, errors='ignore') as file: # Turkish
            filedata = file.read()
        file.close()
    print('Loaded treatment file:' + filepath)
    print('')
    return(filedata)

def save_treatment_file(treatment_text, filepath, encoding = "UTF-8"):
    #with open(filepath, 'w', encoding="ISO-8859-1") as file: 
    try:
        with open(filepath, 'w', encoding=encoding) as file:
            file.write(treatment_text)
        file.close()
    except:
        print('! Warning: Non writable characters exist for selected encoding ' + encoding + '. \n Ignoring non convertible chars.\n Please check if you have selected right encoding')
        with open(filepath, 'w', encoding=encoding, errors='ignore') as file:
            file.write(treatment_text)
        file.close()
    print('Saved treatment file:' + filepath)


def get_matched_entries(textblock):
    matched_items = re.findall("\[\[.*?\]\]", textblock)
    if not matched_items:
#        print('no keywords exist')
        return(False)
    unique_matched_items = list(set(matched_items))
    unique_matched_items.sort()
    print('    Grabbed following keywords:')
    print('  ' + '-' * 35)
    for item in unique_matched_items: print('    ' + item)
    print('-' * 35)
    print(str(len(unique_matched_items)) + ' items in total')
    print('')
    return(unique_matched_items)

def create_own_list(keywordlist):
    try:
        stripped_keywordlist = list(map(lambda each:each.replace('[[','').replace(']]',''), keywordlist))
        return([keywordlist,stripped_keywordlist])
    except:
        if keywordlist == False:
            print('!!Error: Some problem occurred. This might be due to an empty list of keywords.')
        return(keywordlist)
def list_to_dict(keywordlist):
    return(dict(zip(keywordlist[0],keywordlist[1])))

# Retired function below
# def list_to_csv(keywordlist, savepath):
    # try:
        # with open(savepath, 'w', encoding="UTF-8", newline='') as myfile:
            # writer = csv.writer(myfile, delimiter=';', quotechar = '"')
            # seperator_tag = "sep=;"
            # writer.writerow([seperator_tag])
            # for row1,row2 in zip(keywordlist[0], keywordlist[1]):
                # writer.writerow([row1, row2])
                
        # myfile.close()
        # print('succesfully saved to ' + savepath)
    # except:
        # print('An error occured while saving the file')

def list_to_xlsx(keywordlist, savepath):
    try:
        workbook = xlsxwriter.Workbook(savepath)
        worksheet = workbook.add_worksheet()
        rowno = 0
        for col1, col2 in zip(keywordlist[0], keywordlist[1]):
            worksheet.write(rowno, 0, col1)
            worksheet.write(rowno, 1, col2)
            rowno += 1
        workbook.close()
        print('Succesfully saved to ' + savepath)
        print('')
    except:
        print('!! Error: A problem occured while saving the file')



# # Retired function below
# def csv_to_dictionary(filepath):
    # pattern = re.compile("\[\[.*?\]\]")
    
    # language_dict = dict()
    # with open(filepath) as csvfile:
        # reader = csv.reader(csvfile,delimiter=';')
        # for row in reader:
            # if (len(row) >1):
                # if pattern.match(row[0]):
                    # language_dict[row[0]]=row[1]
    # csvfile.close()
    # print('Loaded language file:' + filepath)
    # print('    Grabbed following keywords:')
    # for item in language_dict: print("---- \n     " + item + "\n" + language_dict[item])
    # print('-' * 10)
    # print(str(len(language_dict)) + ' items in total')
    # return(language_dict)


def xlsx_to_dictionary(filepath):
    try:
        pattern = re.compile("\[\[.*?\]\]")

        language_dict = dict()
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active

        for row in worksheet.iter_rows():
            if (len(row) > 1):
                if pattern.match(row[0].value):
                    language_dict[row[0].value] = row[1].value

        workbook.close()
        print('Loaded language file:' + filepath)
        print('')
        print('    Grabbed following translations:')
        print('  ' + '-' * 35)
        for item in language_dict: print(" \n     " + item + "\n    -> " + language_dict[item])
        if (not len(language_dict) > 0):
            print("No keywords exist in the dictionary file")
            return
        print('  ' + '-' * 35)
        print(str(len(language_dict)) + ' items in total')
        return (language_dict)
    except:
        print("Error: A problem occured with reading excel file. Does it contain the keys as the first column and the text in second column?")
        return


def replace_from_dictionary(language_dict, sourcefile):
    new_sourcefile = sourcefile
    for key in language_dict:
        new_sourcefile = new_sourcefile.replace(key, language_dict[key])
        
    return(new_sourcefile)

    
### Function below will be only used for future implementation of the command line functionality
# def generate_language_file(path_treatment_in, path_language_out):
    # source_text = load_treatment_file(path_treatment_in)
    # matched_entries = get_matched_entries(source_text)
    # if not matched_entries:
        # print('no file has been generated')
        # return(False)
    # language_list = create_own_list(matched_entries)
    # list_to_csv(language_list, path_language_out)
        

def implement_language_file(path_treatment_in, path_language_in, path_treatment_out, strip_unmatched,encoding_input = 'UTF-8',encoding_output = 'UTF-8'):
    source_text = load_treatment_file(path_treatment_in, encoding_input)

    # print('  ' + '-' * 20)
    print(' ' + '======== Treatment file ========')
    matched_entries = get_matched_entries(source_text)
    if not matched_entries:
        print('! Error: No matched entries in treatment file. No file is generated')
        return(False)
        
    print('')
    print(' ' + '======== Language file ========')
    language_dict = xlsx_to_dictionary(path_language_in)
    if (not language_dict):
        print('Error with the language file. No treatment file has been generated.')
        return
    set_treatment = set(matched_entries)
    set_language = set(language_dict.keys())
    non_replaced_keys = set_treatment.difference(set_language)
    non_used_keys = set_language.difference(set_treatment)
    
    if non_replaced_keys:
        print('\n')
        print('! Non replaced keys in the treatment file:')
        for item in non_replaced_keys: print(' ' + item,end =",")
        print('')
        print('-' * 10)
        
    if non_used_keys:
        print('\n')
        print('!  Non used keys:')
        for item in non_used_keys: print(' ' + item,end =",")
        print('')
        print('-' * 10)

    if non_used_keys:
        print('! Warning: ' + str(len(non_used_keys)) + ' dictionary items were not used')
    
    if non_replaced_keys:
        print('')
        print('! Warning: ' + str(len(non_replaced_keys)) + ' keys were not replaced by a dictionary input')
        if strip_unmatched:
            print('        stripping brackets from non-replaced keys')
    
    target_text = replace_from_dictionary(language_dict, source_text)
    if strip_unmatched:
        target_text = strip_brackets(target_text)
        printed('stipped brackets non-replaced keys succesfully') 
    
    save_treatment_file(target_text, path_treatment_out,encoding_output)

def strip_brackets(source_text):
    matched_entries = get_matched_entries(source_text)
    language_list = create_own_list(matched_entries)
    language_dict = list_to_dict(language_list)
    target_text = replace_from_dictionary(language_dict, source_text)
    return(target_text)

def strip_brackets_file(path_treatment_in, path_treatment_out,encoding='UTF-8'):
    source_text = load_treatment_file(path_treatment_in,encoding)
    target_text = strip_brackets(source_text)
#    matched_entries = get_matched_entries(source_text)
#    language_list = create_own_list(matched_entries)
#    language_dict = list_to_dict(language_list)
#    target_text = replace_from_dictionary(language_dict, source_text)
    save_treatment_file(target_text, path_treatment_out, encoding)
        
#def save_file(target):
#    #print("This function functions")
#    #QMessageBox.about(self,"Example", "I am clicked")
#    filepath = QFileDialog.getSaveFileName(self, "Save File",
#                           "c:\\",
#                           "Text files (*.txt);;All files (*.*)");
#
#


#
#
#####) OLD FUNCTIONS ########################################
#def create_own_dictionary(keywordlist):
#    ## TODO: Add brackets in the keyword column
#    stripped_keywordlist = list(map(lambda each:each.replace('[[','').replace(']]',''), keywordlist))
#    return({'keyword':keywordlist,'text':stripped_keywordlist})
#
#
#def dictionary_to_file(key_dict, savepath, type = 'csv'):
#    initial_df = pandas.DataFrame(data=key_dict)
#    if (type == 'csv'):
#        initial_df.to_csv(savepath,index=False, header=False)
#    elif (type == 'xls'):
#        initial_df.to_excel(savepath,index=False, header=False)
#    else:
#        raise ValueError('Wrong file format')
#    
#    print('succesfully saved to ' + savepath)
#    return(savepath)
#    
#
#
#def file_to_dictionary(filepath, type = 'csv'):
#    if (type == 'csv'):
#        language_df = pandas.read_csv(filepath, header = None)
#    elif(type == 'xls'):
#        language_df = pandas.read_excel(filepath, header = None)
#    else:
#        raise ValueError('Wrong file format')
#    
#    language_dict = dict()
#    for index, row in language_df.iterrows():
#        language_dict[row[0]]=row[1]
#    
#    return(language_dict)
#############################################################

    

#%%
#    
##%%
## ----------------------
#input_filename = r"C:\Users\saral\ownCloud\decisionlab\zBrac\examples\german-chars\germanchars.txt"
##path_treatment_in = r"C:\zbractest\pd_new_export.txt"
##path_language_in = r"C:/zbractest/ultimatum_tr.csv"

#empty_input_filename = r"C:\zbractest\empty.txt"
##savepath = r'C:\Users\anna\ownCloud\decisionlab\zBrac\sample-files\pd-output3.csv'
##savepath2 = r'C:\Users\anna\ownCloud\decisionlab\zBrac\sample-files\excel.xls'
#xlsx_output = r"C:\Users\saral\ownCloud\decisionlab\zBrac\examples\german-chars\xlsout.xlsx"
#dictionary_input = r"C:\zbractest\pd_dict-eng.csv"
#output_filename = r"C:\zbractest\pd_dict-edited.txt"
#textfile = load_treatment_file(input_filename)
#
#
#matched_entries = get_matched_entries(textfile)
#
#initial_list = create_own_list(matched_entries)
#
#list_to_csv(initial_list, csv_output)
#
#
#language_dict = csv_to_dictionary(dictionary_input)
#
#new_textfile =  replace_from_dictionary(language_dict, textfile)
#
#save_source_file(new_textfile, output_filename)
#


